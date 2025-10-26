"""
Utility functions for Contact export
"""
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def export_contacts_to_excel(contacts, selected_fields=None):
    """
    Export contacts to Excel file
    Returns base64 encoded string of Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    
    # Define all possible fields
    if selected_fields:
        fields = selected_fields
    else:
        fields = [
            'first_name', 'last_name', 'title', 'email', 'email_status',
            'company', 'industry', 'employees', 'annual_revenue',
            'city', 'state', 'country', 'technologies', 'keywords',
            'person_linkedin_url', 'website'
        ]
    
    # Header row with styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for idx, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=idx, value=field.replace('_', ' ').title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_idx, contact in enumerate(contacts, 2):
        for col_idx, field in enumerate(fields, 1):
            value = getattr(contact, field, '')
            if value is None:
                value = ''
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for col_idx, field in enumerate(fields, 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Encode to base64
    excel_data = output.read()
    base64_encoded = base64.b64encode(excel_data).decode('utf-8')
    
    return base64_encoded


def export_contacts_to_csv(contacts, selected_fields=None):
    """
    Export contacts to CSV
    Returns CSV string
    """
    import csv
    
    if selected_fields is None:
        fields = [
            'first_name', 'last_name', 'title', 'email', 'company',
            'industry', 'city', 'country', 'employees'
        ]
    else:
        fields = selected_fields
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fields)
    writer.writeheader()
    
    for contact in contacts:
        row = {field: str(getattr(contact, field, '')) for field in fields}
        writer.writerow(row)
    
    return output.getvalue()

